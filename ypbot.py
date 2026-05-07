import os
import json
import logging
from datetime import datetime
from io import BytesIO

from telegram import Update, ReplyKeyboardRemove, InlineKeyboardButton, InlineKeyboardMarkup
from telegram.ext import (
    Application, CommandHandler, MessageHandler,
    ConversationHandler, CallbackQueryHandler, filters, ContextTypes
)
from telegram.request import HTTPXRequest
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

logging.basicConfig(
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
    level=logging.INFO
)
logger = logging.getLogger(__name__)

# ─── SOZLAMALAR ───────────────────────────────────────────────
BOT_TOKEN = "YOUR_BOT_TOKEN_HERE"
ADMIN_IDS = [123456789]
DB_FILE   = "users.json"
# ──────────────────────────────────────────────────────────────

DISTRICTS = {
    "01": "Toshkent sh.",      "02": "Toshkent v.",
    "03": "Andijon",           "04": "Farg'ona",
    "05": "Namangan",          "06": "Samarqand",
    "07": "Buxoro",            "08": "Navoiy",
    "09": "Qashqadaryo",       "10": "Surxondaryo",
    "11": "Jizzax",            "12": "Sirdaryo",
    "13": "Xorazm",            "14": "Qoraqalpog'iston",
}

COMMITTEES = {
    "k01": "Huquqiy savodxonlik va vatanparvarlik",
    "k02": "Bandlik, sog'liqni saqlash va ijtimoiy masalalar",
    "k03": "Innovatsion rivojlanish, sun'iy intellekt va IT",
    "k04": "Tadbirkorlik, raqobat va sanoat",
    "k05": "Fan, ta'lim, madaniyat, turizm va sport",
    "k06": "Ekologiya va atrof-muhitni muhofaza qilish",
    "k07": "Xalqaro ishlar va yoshlar tashkilotlari",
    "k08": "Fuqarolik jamiyati va volontyorlik",
}

# To'liq rasmiy nomlar (Excel va tasdiqlash xabarida ishlatiladi)
COMMITTEES_FULL = {
    "k01": "Yoshlarning huquqiy savodxonligini oshirish va vatanparvarlik ruhida tarbiyalash masalalari qo'mitasi",
    "k02": "Yoshlar bandligi, sog'liqni saqlash va ijtimoiy masalalar qo'mitasi",
    "k03": "Innovatsion rivojlanish, sun'iy intellekt va axborot texnologiyalari masalalari qo'mitasi",
    "k04": "Yoshlar tadbirkorligi, raqobatni rivojlantirish va sanoat masalalari qo'mitasi",
    "k05": "Fan, ta'lim, madaniyat, turizm va sport masalalari qo'mitasi",
    "k06": "Ekologiya va atrof-muhitni muhofaza qilish masalalari qo'mitasi",
    "k07": "Xalqaro ishlar va yoshlar tashkilotlari bilan hamkorlik qilish masalalari qo'mitasi",
    "k08": "Fuqarolik jamiyati va volontyorlik faoliyatini rivojlantirish masalalari qo'mitasi",
}

POSITIONS = {
    "p01": "Yoshlar parlamenti raisi o'rinbosari",
    "p02": "Yoshlar parlamenti Matbuot kotibi",
    "p03": "Yoshlar parlamenti Ijrochi kotibi",
    "p04": "Qo'mita raisi",
    "p05": "Qo'mita raisi o'rinbosari",
    "p06": "Qo'mita a'zosi",
}

# ConversationHandler holatlari
NAME, DISTRICT, AGE, PHONE, COMMITTEE, POSITION = range(6)


# ─── DATABASE ─────────────────────────────────────────────────
def load_db() -> list:
    if not os.path.exists(DB_FILE):
        return []
    with open(DB_FILE, "r", encoding="utf-8") as f:
        return json.load(f)

def save_db(data: list):
    with open(DB_FILE, "w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

def next_id() -> str:
    return str(len(load_db()) + 1).zfill(4)

def add_user(record: dict):
    data = load_db()
    data.append(record)
    save_db(data)

def already_registered(telegram_id: int) -> bool:
    return any(u.get("TelegramID") == telegram_id for u in load_db())


# ─── KLAVIATURALAR ────────────────────────────────────────────
def district_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    items = list(DISTRICTS.items())
    for i in range(0, len(items), 2):
        row = []
        for code, name in items[i:i+2]:
            row.append(InlineKeyboardButton(f"{code}. {name}", callback_data=f"dist_{code}"))
        buttons.append(row)
    return InlineKeyboardMarkup(buttons)

def committee_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    for code, name in COMMITTEES.items():
        buttons.append([InlineKeyboardButton(name, callback_data=f"com_{code}")])
    return InlineKeyboardMarkup(buttons)

def position_keyboard() -> InlineKeyboardMarkup:
    buttons = []
    for code, name in POSITIONS.items():
        buttons.append([InlineKeyboardButton(name, callback_data=f"pos_{code}")])
    return InlineKeyboardMarkup(buttons)


# ─── EXCEL YARATISH ───────────────────────────────────────────
def build_excel() -> BytesIO:
    users = load_db()
    wb = openpyxl.Workbook()

    ws = wb.active
    ws.title = "Foydalanuvchilar"

    headers    = ["ID", "To'liq ism", "Hudud", "Yosh", "Telefon", "Rol", "Reyting", "Qo'mita", "Lavozim"]
    col_widths = [7, 24, 18, 6, 17, 8, 10, 52, 38]

    header_fill = PatternFill("solid", fgColor="1D4E8F")
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=11)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    left   = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    thin   = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"),  bottom=Side(style="thin")
    )

    for col, (h, w) in enumerate(zip(headers, col_widths), 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center
        cell.border    = thin
        ws.column_dimensions[cell.column_letter].width = w

    ws.row_dimensions[1].height = 24

    even_fill = PatternFill("solid", fgColor="E8EEF7")
    for row_i, u in enumerate(users, 2):
        row_data = [
            u.get("ID"),
            u.get("FullName"),
            DISTRICTS.get(u.get("District", ""), u.get("District", "")),
            u.get("Age"),
            u.get("Phone", ""),
            u.get("Role"),
            u.get("InitialRating"),
            u.get("Committee", ""),
            u.get("Lavozim", ""),
        ]
        fill = even_fill if row_i % 2 == 0 else None
        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_i, column=col, value=val)
            cell.font      = Font(name="Arial", size=10)
            cell.alignment = center if col in (1, 4, 6, 7) else left
            cell.border    = thin
            if fill:
                cell.fill = fill
        ws.row_dimensions[row_i].height = 20

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:I{max(len(users) + 1, 2)}"

    # Ko'rsatmalar varaqasi
    ws2 = wb.create_sheet("Ko'rsatmalar")
    instructions = [
        ("Maydon",      "Tavsif",                             "Namuna"),
        ("ID",          "4 xonali raqam",                     "0001"),
        ("To'liq ism",  "Familiya va ismi",                   "Valiyev Ali"),
        ("Hudud",       "Viloyat yoki shahar nomi",           "Toshkent sh."),
        ("Yosh",        "18 dan 35 gacha",                    "22"),
        ("Telefon",     "Ixtiyoriy",                          "+998901234567"),
        ("Rol",         "user / admin",                       "user"),
        ("Reyting",     "Boshlang'ich ball (0-20)",           "10"),
        ("Qo'mita",     "To'liq qo'mita nomi",               "Fan, ta'lim... qo'mitasi"),
        ("Lavozim",     "Tanlangan lavozim",                  "Qo'mita a'zosi"),
    ]
    for r, row in enumerate(instructions, 1):
        for c, val in enumerate(row, 1):
            cell = ws2.cell(row=r, column=c, value=val)
            cell.font      = Font(bold=(r == 1), name="Arial", size=10,
                                  color="FFFFFF" if r == 1 else "000000")
            cell.fill      = PatternFill("solid", fgColor="1D4E8F") if r == 1 else PatternFill("solid", fgColor="FFFFFF")
            cell.border    = thin
            cell.alignment = left
    ws2.column_dimensions["A"].width = 14
    ws2.column_dimensions["B"].width = 34
    ws2.column_dimensions["C"].width = 24

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


# ─── HANDLERLAR ───────────────────────────────────────────────
async def start(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    user = update.effective_user
    if already_registered(user.id):
        await update.message.reply_text(
            "Siz allaqachon ro'yxatdan o'tgansiz.\n\n"
            "Yordam uchun /help buyrug'ini yuboring."
        )
        return ConversationHandler.END

    await update.message.reply_text(
        f"Assalomu alaykum, {user.first_name}.\n\n"
        "Ro'yxatdan o'tish uchun quyidagi savollarga javob bering.\n"
        "Bekor qilish: /cancel\n\n"
        "Ismingiz va familiyangizni kiriting:"
    )
    return NAME


async def get_name(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    name = update.message.text.strip()
    if len(name) < 3:
        await update.message.reply_text(
            "To'liq ism familiyangizni kiriting (kamida 3 ta belgi)."
        )
        return NAME

    context.user_data["name"] = name
    await update.message.reply_text(
        f"Qabul qilindi: {name}\n\nYashash hududingizni tanlang:",
        reply_markup=district_keyboard()
    )
    return DISTRICT


async def get_district(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    code = query.data.replace("dist_", "")
    name = DISTRICTS.get(code, code)
    context.user_data["district"] = code

    await query.edit_message_text(
        f"Hudud tanlandi: {code}. {name}\n\n"
        "Yoshingizni kiriting (18 dan 35 gacha):"
    )
    return AGE


async def get_age(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    try:
        age = int(update.message.text.strip())
        if not (18 <= age <= 35):
            raise ValueError
    except ValueError:
        await update.message.reply_text("Yosh 18 dan 35 oralig'ida bo'lishi kerak. Qaytadan kiriting:")
        return AGE

    context.user_data["age"] = age
    await update.message.reply_text(
        f"Qabul qilindi: {age} yosh\n\n"
        "Telefon raqamingizni kiriting.\n"
        "Ixtiyoriy — o'tkazib yuborish uchun «-» belgi yozing:"
    )
    return PHONE


async def get_phone(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    text = update.message.text.strip()
    phone = "" if text in ("-", "–", "yo'q", "skip") else text
    context.user_data["phone"] = phone

    await update.message.reply_text(
        f"Qabul qilindi: {phone or 'ko\'rsatilmagan'}\n\n"
        "Qo'mitangizni tanlang:",
        reply_markup=committee_keyboard()
    )
    return COMMITTEE


async def get_committee(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    code = query.data.replace("com_", "")
    short_name = COMMITTEES.get(code, code)
    full_name  = COMMITTEES_FULL.get(code, code)
    context.user_data["committee_code"] = code
    context.user_data["committee_full"] = full_name

    await query.edit_message_text(
        f"Qo'mita tanlandi:\n{full_name}\n\n"
        "Lavozimingizni tanlang:",
        reply_markup=position_keyboard()
    )
    return POSITION


async def get_position(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    query = update.callback_query
    await query.answer()

    code     = query.data.replace("pos_", "")
    position = POSITIONS.get(code, code)
    context.user_data["position"] = position

    uid       = next_id()
    dist_code = context.user_data["district"]
    dist_name = DISTRICTS.get(dist_code, dist_code)
    phone     = context.user_data["phone"]
    committee = context.user_data["committee_full"]

    record = {
        "ID":            uid,
        "FullName":      context.user_data["name"],
        "District":      dist_code,
        "Age":           context.user_data["age"],
        "Phone":         phone,
        "Role":          "user",
        "InitialRating": 10,
        "Committee":     committee,
        "Lavozim":       position,
        "TelegramID":    update.effective_user.id,
        "RegisteredAt":  datetime.now().strftime("%Y-%m-%d %H:%M"),
    }
    add_user(record)

    await query.edit_message_text(
        f"Ro'yxatdan o'tish muvaffaqiyatli yakunlandi.\n\n"
        f"ID: {uid}\n"
        f"Ism: {record['FullName']}\n"
        f"Hudud: {dist_code}. {dist_name}\n"
        f"Yosh: {record['Age']}\n"
        f"Telefon: {phone or 'ko\'rsatilmagan'}\n"
        f"Reyting: {record['InitialRating']}\n"
        f"Qo'mita: {committee}\n"
        f"Lavozim: {position}\n\n"
        "Ma'lumotlaringiz saqlandi. Rahmat."
    )
    logger.info(f"Yangi a'zo: {record['FullName']} (ID: {uid}), {position}")
    return ConversationHandler.END


async def cancel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> int:
    context.user_data.clear()
    await update.message.reply_text(
        "Ro'yxatdan o'tish bekor qilindi.\n"
        "Qaytadan boshlash uchun /start buyrug'ini yuboring.",
        reply_markup=ReplyKeyboardRemove()
    )
    return ConversationHandler.END


# ─── ADMIN HANDLERLAR ─────────────────────────────────────────
def is_admin(update: Update) -> bool:
    return update.effective_user.id in ADMIN_IDS


async def admin_stats(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("Ruxsat yo'q.")
        return

    users = load_db()
    total   = len(users)
    w_phone = sum(1 for u in users if u.get("Phone"))

    by_dist = {}
    for u in users:
        d = u.get("District", "?")
        by_dist[d] = by_dist.get(d, 0) + 1

    by_com = {}
    for u in users:
        c = u.get("Committee", "Noma'lum")
        by_com[c] = by_com.get(c, 0) + 1

    top_dist = sorted(by_dist.items(), key=lambda x: -x[1])[:5]
    dist_txt = "\n".join(f"  {DISTRICTS.get(c, c)}: {n} ta" for c, n in top_dist) or "  —"

    top_com = sorted(by_com.items(), key=lambda x: -x[1])[:3]
    com_txt = "\n".join(f"  {c[:40]}...: {n} ta" if len(c) > 40 else f"  {c}: {n} ta"
                        for c, n in top_com) or "  —"

    await update.message.reply_text(
        f"Statistika\n\n"
        f"Jami a'zolar: {total} ta\n"
        f"Telefon ko'rsatganlar: {w_phone} ta\n\n"
        f"Top hududlar:\n{dist_txt}\n\n"
        f"Eng ko'p qo'mitalar:\n{com_txt}"
    )


async def admin_excel(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("Ruxsat yo'q.")
        return

    users = load_db()
    if not users:
        await update.message.reply_text("Hali ro'yxatdan o'tgan a'zo yo'q.")
        return

    msg = await update.message.reply_text("Excel fayl tayyorlanmoqda...")
    buf      = build_excel()
    filename = f"azolar_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"

    await update.message.reply_document(
        document=buf,
        filename=filename,
        caption=f"Jami: {len(users)} ta a'zo | {datetime.now().strftime('%d.%m.%Y %H:%M')}"
    )
    await msg.delete()


async def user_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    await update.message.reply_text(
        "Yordam\n\n"
        "/start  — Ro'yxatdan o'tish\n"
        "/cancel — Bekor qilish"
    )


async def admin_help(update: Update, context: ContextTypes.DEFAULT_TYPE):
    if not is_admin(update):
        await update.message.reply_text("Ruxsat yo'q.")
        return
    await update.message.reply_text(
        "Admin buyruqlar\n\n"
        "/stats — Statistika\n"
        "/excel — Excel yuklab olish\n"
        "/help  — Yordam"
    )


# ─── MAIN ─────────────────────────────────────────────────────
def main():
    PROXY_URL = None  # kerak bo'lsa: "http://127.0.0.1:8080"

    request = HTTPXRequest(
        connect_timeout=30.0,
        read_timeout=30.0,
        write_timeout=30.0,
        pool_timeout=30.0,
        proxy=PROXY_URL,
    )

    app = (
        Application.builder()
        .token(BOT_TOKEN)
        .request(request)
        .build()
    )

    conv = ConversationHandler(
        entry_points=[CommandHandler("start", start)],
        states={
            NAME:      [MessageHandler(filters.TEXT & ~filters.COMMAND, get_name)],
            DISTRICT:  [CallbackQueryHandler(get_district,  pattern=r"^dist_")],
            AGE:       [MessageHandler(filters.TEXT & ~filters.COMMAND, get_age)],
            PHONE:     [MessageHandler(filters.TEXT & ~filters.COMMAND, get_phone)],
            COMMITTEE: [CallbackQueryHandler(get_committee, pattern=r"^com_")],
            POSITION:  [CallbackQueryHandler(get_position,  pattern=r"^pos_")],
        },
        fallbacks=[CommandHandler("cancel", cancel)],
    )

    app.add_handler(conv)
    app.add_handler(CommandHandler("stats", admin_stats))
    app.add_handler(CommandHandler("excel", admin_excel))
    app.add_handler(CommandHandler("help",  user_help))

    logger.info("Bot ishga tushdi.")
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
